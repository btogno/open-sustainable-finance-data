"""
import_corpus.py — convert the hand-coded workbook into the frozen corpus CSV.

Usage:  python3 scripts/import_corpus.py <workbook.xlsx> [--sheet "Full Coding"]

The workbook is the coding instrument: it carries a merged group-header band in
row 1, the real column names in row 2, and one paper per row thereafter. This
script normalises the header names to the canonical CSV schema (stripping the
line breaks and the "★" markers the workbook uses for legibility), verifies the
column order has not moved, and writes data/SustFin_Corpus_FINAL.csv.

It refuses to write if the column set does not match the canonical schema, so a
reordered or renamed workbook column cannot silently corrupt the corpus.
"""

import csv
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEST = os.path.join(ROOT, "data", "SustFin_Corpus_FINAL.csv")

CANONICAL = [
    "PAPER ID",
    "Journal",
    "Full Citation (Author, Year, Journal)",
    "Short Title/ Keyword",
    "Description of data",
    "Direct link",
    "Topic: Climate Physical Risk",
    "Topic: Climate Transition Risk",
    "Topic: ESG Disclosure",
    "Topic: ESG Ratings",
    "Topic: Biodiversity & Nature",
    "Topic: Green Bonds & Sustainable Debt",
    "Topic: Corporate Emissions (Scope 1/2/3)",
    "Topic: Social Factors",
    "Topic: Governance",
    "Topic: Other",
    "Method (primary)",
    "Method (secondary)",
    "Method (if other)",
    "Data Source (primary)",
    "Data Source (secondary)",
    "Data Source (if other)",
    "Publication Year",
    "Dataset Start Year",
    "Dataset End Year",
    "Data Geographic Scope",
    "Geographic Notes",
    "Data Publicly Available? (Y/N/Partial)",
    "Data Access Link / Repository",
    "Replication Code? (Y/N/Partial)",
    "Code Access Link / Repository",
    "Could be open-source (Y/N/Partial)",
    "Unit of Observation",
]

# Workbook header -> canonical header, where the two differ beyond whitespace.
ALIASES = {
    "Data Access Link": "Data Access Link / Repository",
    "Code Access Link": "Code Access Link / Repository",
}


def norm(h):
    """Normalise a workbook header: drop line breaks, stars and stray spacing."""
    h = re.sub(r"[★*]", " ", str(h or ""))
    h = re.sub(r"\s+", " ", h.replace("\n", " ")).strip()
    return ALIASES.get(h, h)


def cell(v):
    """Render a cell as the CSV stores it: text, with years as bare integers."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def read_workbook(path, sheet="Full Coding"):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"sheet {sheet!r} not found; workbook has {wb.sheetnames}")
    rows = list(wb[sheet].iter_rows(values_only=True))

    # Find the header row: the first row whose first cell normalises to PAPER ID.
    hdr_ix = next((i for i, r in enumerate(rows) if norm(r[0]) == "PAPER ID"), None)
    if hdr_ix is None:
        sys.exit("could not find a header row containing 'PAPER ID'")

    headers = [norm(h) for h in rows[hdr_ix]]
    if headers != CANONICAL:
        diff = [
            f"    col {i}: workbook {got!r} != canonical {want!r}"
            for i, (got, want) in enumerate(zip(headers, CANONICAL))
            if got != want
        ]
        extra = len(headers) - len(CANONICAL)
        sys.exit(
            "workbook columns do not match the canonical schema; refusing to write.\n"
            + "\n".join(diff)
            + (f"\n    column count differs by {extra}" if extra else "")
        )

    data = [r for r in rows[hdr_ix + 1:] if cell(r[0])]
    return headers, data


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__.strip())
    src = sys.argv[1]
    sheet = "Full Coding"
    if "--sheet" in sys.argv:
        sheet = sys.argv[sys.argv.index("--sheet") + 1]

    headers, data = read_workbook(src, sheet)
    ids = [cell(r[0]) for r in data]
    if len(set(ids)) != len(ids):
        dupes = sorted({i for i in ids if ids.count(i) > 1})
        sys.exit(f"duplicate paper ids in the workbook: {dupes}")

    with open(DEST, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(headers)
        for r in data:
            w.writerow([cell(v) for v in list(r)[: len(headers)]])

    print(f"wrote {os.path.relpath(DEST, ROOT)} — {len(data)} rows from sheet {sheet!r}")


if __name__ == "__main__":
    main()
